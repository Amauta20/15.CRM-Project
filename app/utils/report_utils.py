import openpyxl
from openpyxl.styles import Font, PatternFill
import datetime
from app.utils import time_utils
from app.opportunities.opportunities_manager import get_opportunity_by_id
from app.data.contacts_manager import get_contact_by_id

def generate_excel_report(report_data, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Kanban Report"

    # Define headers
    headers = ["ID", "Título", "Descripción", "Columna", "Fecha de Vencimiento", "Asignado a", "Creado el", "Iniciada", "Finalizada"]
    sheet.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_idx, cell in enumerate(sheet[1]):
        cell.font = header_font
        cell.fill = header_fill
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = 20 # Default width

    # Write data
    for card in report_data:
        due_date_formatted = ""
        if card.get("due_date"):
            try:
                dt_obj = datetime.datetime.fromisoformat(card["due_date"].replace("Z", "+00:00"))
                due_date_formatted = time_utils.format_datetime(time_utils.from_utc(dt_obj))
            except (ValueError, TypeError):
                due_date_formatted = card["due_date"]

        created_at_formatted = ""
        if card.get("created_at"):
            try:
                dt_obj = datetime.datetime.fromisoformat(card["created_at"].replace("Z", "+00:00"))
                created_at_formatted = time_utils.format_datetime(time_utils.from_utc(dt_obj))
            except (ValueError, TypeError):
                created_at_formatted = card["created_at"]

        started_at_formatted = ""
        if card.get("started_at"):
            try:
                dt_obj = datetime.datetime.fromisoformat(card["started_at"].replace("Z", "+00:00"))
                started_at_formatted = time_utils.format_datetime(time_utils.from_utc(dt_obj))
            except (ValueError, TypeError):
                started_at_formatted = card["started_at"]

        finished_at_formatted = ""
        if card.get("finished_at"):
            try:
                dt_obj = datetime.datetime.fromisoformat(card["finished_at"].replace("Z", "+00:00"))
                finished_at_formatted = time_utils.format_datetime(time_utils.from_utc(dt_obj))
            except (ValueError, TypeError):
                finished_at_formatted = card["finished_at"]

        row_data = [
            card.get("id", ""),
            card.get("title", ""),
            card.get("description", ""),
            card.get("column_name", ""),
            due_date_formatted,
            card.get("assigned_to", ""),
            created_at_formatted,
            started_at_formatted,
            finished_at_formatted
        ]
        sheet.append(row_data)

    # Adjust column widths based on content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(file_path)

def generate_proposal_document(opportunity_id, file_path):
    opportunity = get_opportunity_by_id(opportunity_id)
    if not opportunity:
        return False

    contact = None
    if opportunity['contact_id']:
        contact = get_contact_by_id(opportunity['contact_id'])

    with open(file_path, "w", encoding="utf-8") as f:
        f.write(f"--- Propuesta para la Oportunidad: {opportunity['title']} ---\n\n")
        f.write(f"Requerimiento: {opportunity['requirement']}\n")
        f.write(f"Monto: {opportunity['amount']} {opportunity['currency'] or 'USD'}\n")
        f.write(f"Fase: {opportunity['phase']}\n")
        f.write(f"Fecha de Entrega: {opportunity['delivery_date']}\n")
        f.write(f"Probabilidad de Éxito: {opportunity['success_probability']}%\n")
        f.write(f"Creado el: {time_utils.format_datetime(time_utils.from_utc(datetime.datetime.fromisoformat(opportunity['created_at'].replace('Z', '+00:00'))))}\n")
        f.write(f"Última Actualización: {time_utils.format_datetime(time_utils.from_utc(datetime.datetime.fromisoformat(opportunity['updated_at'].replace('Z', '+00:00'))))}\n")

        if contact:
            f.write(f"\n--- Información de Contacto ---\n")
            f.write(f"Nombre: {contact['name']}\n")
            f.write(f"Email: {contact['email'] or 'N/A'}\n")
            f.write(f"Teléfono: {contact['phone'] or 'N/A'}\n")

        f.write(f"\n---------------------------------------\n")
    return True
